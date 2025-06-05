import json
from openpyxl import load_workbook

def convert_excel_to_json(excel_path, json_path):
    # Load workbook
    wb = load_workbook(excel_path)
    # Get the active worksheet
    ws = wb.active
    
    # Define fixed column indices based on user input
    # A = 0, B = 1, C = 2, D = 3, E = 4, F = 5
    column_indices = {
        "NOMBRES": 2,  # Column C
        "CEDULA": 3,              # Column D
        "CLINICA": 4,             # Column E
        "HOSPITAL": 5             # Column F
    }
    
    # Create list of dictionaries
    data = []
    
    # Process rows
    for row in ws.iter_rows(min_row=2):  # Start from row 2 to skip headers
        row_data = {}
        for header, idx in column_indices.items():
            cell_value = row[idx].value
            # Handle different types of values
            if cell_value is not None:
                if isinstance(cell_value, str):
                    row_data[header] = cell_value.strip()
                elif isinstance(cell_value, (int, float)):
                    row_data[header] = str(cell_value)
                else:
                    row_data[header] = str(cell_value)
        if row_data:  # Only add if we have data
            data.append(row_data)
    
    # Save to JSON file
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    excel_path = "public/LISTADO DE CLINICAS SEMESTRAL PARA PUBLICAR 2025-3.xlsx"
    json_path = "public/clinicas_semestral.json"
    convert_excel_to_json(excel_path, json_path)
    print(f"Conversion completed. JSON file saved as {json_path}")
